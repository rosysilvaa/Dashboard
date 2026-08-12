"""
Lambda de ingestão - Dashboard de Vendas na Nuvem
---------------------------------------------------
Disparada automaticamente por um evento de upload no bucket S3
(ObjectCreated). Lê o arquivo CSV ou Excel, valida e normaliza
cada linha e grava os registros no DynamoDB.

Variáveis de ambiente esperadas:
  TABLE_NAME  -> nome da tabela DynamoDB de vendas
"""

import csv
import io
import os
import uuid
import decimal
import logging
from datetime import datetime, timezone

import boto3

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client("s3")
dynamodb = boto3.resource("dynamodb")

TABLE_NAME = os.environ.get("TABLE_NAME", "vendas")
table = dynamodb.Table(TABLE_NAME)

REQUIRED_COLUMNS = {
    "data", "cliente", "cidade", "uf", "produto",
    "categoria", "quantidade", "valor_unitario", "valor_total", "canal",
}


def _to_decimal(value: str) -> decimal.Decimal:
    """Converte string numérica (com vírgula ou ponto) para Decimal, formato aceito pelo DynamoDB."""
    if value is None or value == "":
        return decimal.Decimal("0")
    cleaned = str(value).replace(".", "").replace(",", ".") if "," in str(value) and "." in str(value) else str(value).replace(",", ".")
    try:
        return decimal.Decimal(cleaned)
    except decimal.InvalidOperation:
        return decimal.Decimal("0")


def _read_csv_bytes(raw_bytes: bytes):
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _read_xlsx_bytes(raw_bytes: bytes):
    """Leitura simples de .xlsx usando openpyxl (precisa estar disponível via layer)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() for h in next(rows_iter)]
    rows = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows


def _validate_row(row: dict) -> bool:
    keys = {k.strip().lower() for k in row.keys()}
    return REQUIRED_COLUMNS.issubset(keys)


def _normalize_row(row: dict) -> dict:
    row = {k.strip().lower(): v for k, v in row.items()}
    quantidade = int(float(str(row["quantidade"]).replace(",", ".")))
    valor_unitario = _to_decimal(row["valor_unitario"])
    valor_total = _to_decimal(row["valor_total"])

    return {
        "id_venda": str(row.get("id_venda") or uuid.uuid4()),
        "data": str(row["data"])[:10],
        "cliente": str(row["cliente"]).strip(),
        "cidade": str(row["cidade"]).strip(),
        "uf": str(row["uf"]).strip().upper(),
        "produto": str(row["produto"]).strip(),
        "categoria": str(row["categoria"]).strip(),
        "quantidade": quantidade,
        "valor_unitario": valor_unitario,
        "valor_total": valor_total,
        "canal": str(row.get("canal", "Não informado")).strip(),
        "ingested_at": datetime.now(timezone.utc).isoformat(),
    }


def handler(event, context):
    processed, failed = 0, 0

    for record in event.get("Records", []):
        bucket = record["s3"]["bucket"]["name"]
        key = record["s3"]["object"]["key"]
        logger.info("Processando s3://%s/%s", bucket, key)

        obj = s3.get_object(Bucket=bucket, Key=key)
        raw_bytes = obj["Body"].read()

        if key.lower().endswith((".xlsx", ".xls")):
            rows = _read_xlsx_bytes(raw_bytes)
        else:
            rows = _read_csv_bytes(raw_bytes)

        with table.batch_writer() as batch:
            for row in rows:
                if not row or not _validate_row(row):
                    failed += 1
                    continue
                try:
                    item = _normalize_row(row)
                    batch.put_item(Item=item)
                    processed += 1
                except Exception as exc:  # noqa: BLE001
                    logger.warning("Falha ao processar linha: %s | erro: %s", row, exc)
                    failed += 1

    logger.info("Ingestão concluída: %d registros gravados, %d falharam", processed, failed)
    return {"processed": processed, "failed": failed}
